#!/usr/bin/python3
import openpyxl, os, pywhatkit, pyautogui, time

wb = openpyxl.load_workbook("contact_list.xlsx")
numbers_sheet = wb.active

archived_nums = []
numbers_to_contact = []

for filename in os.listdir("archive"):
    sheet = openpyxl.load_workbook("archive/" + filename).active
    for i in range(2, sheet.max_row):
        current_number = sheet.cell(row=i, column=2).value
        if not current_number == None and current_number not in archived_nums:
            archived_nums.append(current_number)

for i in range(2, numbers_sheet.max_row):
    num = numbers_sheet.cell(row=i, column=2).value
    was_contacted = numbers_sheet.cell(row=i, column=5).value
    if not num == None and num not in archived_nums and not was_contacted == "x":
        numbers_to_contact.append(num)

for number in numbers_to_contact:
    for i in range(2, numbers_sheet.max_row):
        cell_num = numbers_sheet.cell(row=i, column=2).value
        if number == cell_num:
            print("Sending a text to '" + number + "'....")
            pywhatkit.sendwhatmsg_instantly(number, msg)
            time.sleep(1)
            pyautogui.click()
            pyautogui.hotkey("command", "w")